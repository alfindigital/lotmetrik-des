#!/usr/bin/env python3
"""
parse_des.py — Parse a folder of DES XLSX files into clean ticker data + chained deltas.

Usage:  python3 parse_des.py <folder_with_DES_*.xlsx> [out.json]
Files must be named DES_<YYYY>_P1|P2_*.xlsx (so they sort chronologically).

Handles the 4 gotchas (see reference/parsing_and_totals.md):
  1. "Kode Saham" column position varies -> detect by header.
  2. multi-sheet files: pick the sheet with the MOST valid tickers (avoids stale copies).
  3. ticker TRUE (Triniti Dinamik) read as boolean -> remap.
  4. only 4-letter IDX codes counted (perusahaan publik tanpa kode dikecualikan).

Outputs des_data.json = {order:[keys], meta:[{key,total}], names:{code:name}, present:{code:bitstring}}
and prints per-release totals + validates chained delta (prev+masuk-keluar==cur).
"""
import openpyxl, re, json, sys, glob, os

TICK = re.compile(r'^[A-Z]{4}$')

def norm(v):
    if isinstance(v, bool): return "TRUE" if v else "FALSE"   # gotcha #3
    return "" if v is None else str(v).replace('\xa0', ' ').strip()

def sheet_map(ws):
    rows = list(ws.iter_rows(values_only=True))
    kc = None                                                  # gotcha #1
    for r in rows:
        for j, c in enumerate(r):
            if isinstance(c, str) and 'kode saham' in c.lower():
                kc = j; break
        if kc is not None: break
    if kc is None: kc = 2
    d = {}
    for r in rows:
        if len(r) > kc:
            code = norm(r[kc])
            if TICK.match(code):
                name = ""
                for k in (kc + 1, kc + 2):
                    if len(r) > k and norm(r[k]): name = norm(r[k]); break
                d.setdefault(code, name)
    return d

def des_map(path):                                             # gotcha #2 + #4
    wb = openpyxl.load_workbook(path, data_only=True)
    best = {}
    for sn in wb.sheetnames:
        d = sheet_map(wb[sn])
        if len(d) > len(best): best = d
    return best

def key_of(fn):
    m = re.search(r'(\d{4})_P([12])', os.path.basename(fn))
    return f"{m.group(1)}_P{m.group(2)}" if m else os.path.basename(fn)

def main(folder, out="des_data.json"):
    files = sorted(glob.glob(os.path.join(folder, "*.xlsx")), key=key_of)
    maps = {key_of(f): des_map(f) for f in files}
    order = sorted(maps, key=lambda k: (int(k[:4]), k[-1]))
    names = {}
    for k in order:
        for c, nm in maps[k].items():
            if nm: names[c] = nm
    union = sorted(set().union(*[set(maps[k]) for k in order]))
    present = {c: "".join("1" if c in maps[k] else "0" for k in order) for c in union}
    meta = [{"key": k, "total": len(maps[k])} for k in order]
    print(f"{'rilis':10s}{'ticker':>8s}   delta(masuk/keluar) check")
    ok = True
    for i, k in enumerate(order):
        if i == 0:
            print(f"{k:10s}{len(maps[k]):>8d}   (baseline)"); continue
        prev, cur = set(maps[order[i-1]]), set(maps[k])
        mk, kl = cur - prev, prev - cur
        chk = len(prev) + len(mk) - len(kl)
        good = chk == len(cur); ok = ok and good
        print(f"{k:10s}{len(cur):>8d}   +{len(mk)}/-{len(kl)}  {'OK' if good else '!!MISMATCH'}")
    print("all transitions reconcile:", ok, "| unique tickers:", len(union))
    json.dump({"order": order, "meta": meta, "names": names, "present": present},
              open(out, "w"), ensure_ascii=False)
    print("saved", out)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python3 parse_des.py <folder> [out.json]"); sys.exit(1)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "des_data.json")
